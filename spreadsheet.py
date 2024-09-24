from operator import itemgetter
import openpyxl

spreadsheet_file = openpyxl.load_workbook("employees.xlsx")
employee_spreadsheet = spreadsheet_file["Sheet1"]

employee_spreadsheet.delete_cols(3,4)

employees_by_experience = []

for row in range(2, employee_spreadsheet.max_row + 1):
    employee_name = employee_spreadsheet.cell(row, 1).value
    employee_experience = employee_spreadsheet.cell(row, 2).value

    employees_by_experience.append({
        "name": employee_name,
        "experience": int(employee_experience)
    })

new_list = sorted(employees_by_experience, key=itemgetter("experience"), reverse=True)

for row in range(2, employee_spreadsheet.max_row + 1):
    employee_name = employee_spreadsheet.cell(row, 1)
    employee_experience = employee_spreadsheet.cell(row, 2)

    index = row - 2
    employee = new_list[index]

    employee_name.value = employee["name"]
    employee_experience.value = employee["experience"]

spreadsheet_file.save("employees_sorted_by_experience.xlsx")